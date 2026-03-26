import os
from typing import Dict, List

import numpy
import openpyxl
import pytest
import scadnano as sc


import nuad.constraints as nc
import nuad.search as ns
import nuad.vienna_nupack as nv
from nuad.constraints import (
    BasePairType,
    Design,
    Domain,
    DomainPool,
    Strand,
    StrandDomainAddress,
    _BasePairDomainEndpoint,
    _exterior_base_type_of_domain_3p_end,
    _get_base_pair_domain_endpoints_to_check,
    _get_implicitly_bound_domain_addresses,
)
from nuad.search import Evaluation

_domain_pools: Dict[int, DomainPool] = {}


def assign_domain_pool_of_length(length: int) -> DomainPool:
    """Returns a DomainPool of given size

    :param length: Size of DomainPool
    :type length: int
    :return: DomainPool
    :rtype: DomainPool
    """
    if length in _domain_pools:
        return _domain_pools[length]
    else:
        new_domain_pool = DomainPool(f"POOL_{length}", length)
        _domain_pools[length] = new_domain_pool
        return new_domain_pool


def construct_strand(design: Design, domain_names: List[str], domain_lengths: List[int]) -> Strand:
    """Constructs a strand with given domain names and domain lengths.

    :param domain_names: Names of the domain on the strand
    :type domain_names: List[str]
    :param domain_lengths: Lengths of the domain on the strand
    :type domain_lengths: List[int]
    :raises ValueError: If domain_names is not same size as domain_lengths
    :return: Strand
    :rtype: Strand
    """
    if len(domain_names) != len(domain_lengths):
        raise ValueError(
            f"domain_names and domain_lengths need to contain same"
            f"number of elements but instead found that "
            f"domain_names contained {len(domain_names)} names "
            f"but domain_lengths contained {len(domain_lengths)} "
            f"lengths"
        )
    s: Strand = design.add_strand(domain_names=domain_names)
    for i, length in enumerate(domain_lengths):
        s.domains[i].pool = assign_domain_pool_of_length(length)
    s.compute_derived_fields()
    return s


class TestIntersectingDomains:
    def test_strand_intersecting_domains(self) -> None:
        """
        Test strand construction with nested subdomains

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h
        """
        E = Domain("e", assign_domain_pool_of_length(5), dependent=False)
        F = Domain("f", assign_domain_pool_of_length(5), dependent=False)
        g = Domain("g", assign_domain_pool_of_length(5), dependent=True)
        h = Domain("h", assign_domain_pool_of_length(5), dependent=True)

        b = Domain("b", assign_domain_pool_of_length(10), dependent=True, subdomains=[E, F])
        C = Domain("C", assign_domain_pool_of_length(10), dependent=False, subdomains=[g, h])

        a = Domain("a", assign_domain_pool_of_length(20), dependent=True, subdomains=[b, C])

        all_domains = [a, b, C, E, F, g, h]

        # make strands with different concatenations of domains above that cover the whole tree
        s1 = Strand(domains=[a], starred_domain_indices=[])
        assert len(s1.domains) == 1
        assert s1.domains[0] == a

        s2 = Strand(domains=[b, C], starred_domain_indices=[])
        assert len(s2.domains) == 2
        assert s2.domains[0] == b
        assert s2.domains[1] == C

        s3 = Strand(domains=[E, F, g, h], starred_domain_indices=[])
        assert len(s3.domains) == 4
        assert s3.domains[0] == E
        assert s3.domains[1] == F
        assert s3.domains[2] == g
        assert s3.domains[3] == h

        s4 = Strand(domains=[E, F, C], starred_domain_indices=[])
        assert len(s4.domains) == 3
        assert s4.domains[0] == E
        assert s4.domains[1] == F
        assert s4.domains[2] == C

        for s in [s1, s2, s3, s4]:
            for domain in all_domains:
                assert s.intersects_domain(domain)

        # these strands do not hit every domain
        s5 = Strand(domains=[b, g], starred_domain_indices=[])
        assert len(s5.domains) == 2
        assert s5.domains[0] == b
        assert s5.domains[1] == g

        for domain in [a, b, C, E, F, g]:
            assert s.intersects_domain(domain)
        assert not s5.intersects_domain(h)

        s6 = Strand(domains=[b], starred_domain_indices=[])
        assert len(s6.domains) == 1
        assert s5.domains[0] == b

        for domain in [a, b, E, F]:
            assert s6.intersects_domain(domain)
        for domain in [C, g, h]:
            assert not s6.intersects_domain(domain)


class TestSampleSubstrings:
    def test_substrings(self) -> None:
        sampler = nc.SubstringSampler(supersequence="abcdefghij", substring_length=4, except_start_indices=[2, 3, 5])
        assert sampler.extended_supersequence == "abcdefghij"
        assert sampler.start_indices == (0, 1, 4, 6)

        # sample lots of substrings to ensure we get them all
        rng = numpy.random.default_rng(1)
        substrings = set()
        for _ in range(100):
            substrings.add(sampler.sample_substring(rng))
        substrings = sorted(list(substrings))
        # abcdefghij
        # 0123456789
        #   XX X
        # abcd
        #  bcde
        #     efgh
        #       ghij
        assert substrings == ["abcd", "bcde", "efgh", "ghij"]

    def test_substrings_circular(self) -> None:
        sampler_circular = nc.SubstringSampler(
            supersequence="abcdefghij", substring_length=4, except_start_indices=[1, 3, 5], circular=True
        )
        assert sampler_circular.extended_supersequence == "abcdefghijabc"
        assert sampler_circular.start_indices == (0, 2, 4, 6, 7, 8, 9)

        # sample lots of substrings to ensure we get them all
        rng = numpy.random.default_rng(1)
        substrings = set()
        for _ in range(100):
            substrings.add(sampler_circular.sample_substring(rng))
        substrings = sorted(list(substrings))
        # abcdefghijabc
        # 0123456789012
        #  X X X     X
        # abcd
        #   cdef
        #     efgh
        #       ghij
        #        hija
        #         ijab
        #          jabc
        assert substrings == ["abcd", "cdef", "efgh", "ghij", "hija", "ijab", "jabc"]

    def test_substrings_circular_except_overlapping_indices(self) -> None:
        sampler = nc.SubstringSampler(
            supersequence="abcdefghij", substring_length=3, except_overlapping_indices=[2, 7], circular=True
        )
        assert sampler.extended_supersequence == "abcdefghijab"
        assert sampler.start_indices == (3, 4, 8, 9)

        # sample lots of substrings to ensure we get them all
        rng = numpy.random.default_rng(1)
        substrings = set()
        for _ in range(100):
            substrings.add(sampler.sample_substring(rng))
        substrings = sorted(list(substrings))
        # abcdefghijabc
        # 0123456789
        #   X    X
        #    def
        #     efg
        #         ija
        #          jab
        assert substrings == ["def", "efg", "ija", "jab"]


class TestModifyDesignAfterCreated:
    def setup_method(self) -> None:
        self.design = nc.Design()
        self.design.add_strand(domain_names=["x", "y"])

    def add_domain(self):
        strand = self.design.strands[0]
        strand.domains.append(nc.Domain("z"))
        self.design.compute_derived_fields()
        return strand

    def test_add_domain(self) -> None:
        strand = self.add_domain()

        assert len(self.design.domains) == 3

        actual_domain_names = sorted([d.name for d in self.design.domains])
        assert actual_domain_names == ["x", "y", "z"]

        assert strand.name == "x-y-z"

    def test_add_domain_assign_sequence(self):
        strand = self.add_domain()
        pool = DomainPool("a domain pool", 10)
        for domain in strand.domains:
            domain.pool = pool
        rng = numpy.random.default_rng(0)
        ns.assign_sequences_to_domains_randomly_from_pools(self.design, True, True, rng)

        # assert we don't raise an exception trying to access the sequence of each domain
        s0 = strand.domains[0].sequence  # noqa
        s1 = strand.domains[1].sequence  # noqa
        s2 = strand.domains[2].sequence  # noqa


class TestFromScadnanoDesign:
    def test_two_instances_of_domain(self) -> None:
        """
            x           x
        [--------+ [--------+
                 |          |
        <--------+ <--------+
            y           y*
        """
        helices = [sc.Helix(max_offset=100) for _ in range(2)]
        sc_design = sc.Design(helices=helices)
        sc_design.draw_strand(0, 0).move(10).cross(1).move(-10)
        sc_design.draw_strand(0, 10).move(10).cross(1).move(-10)
        s0, s1 = sc_design.strands
        d00: sc.Domain = s0.domains[0]
        d01: sc.Domain = s0.domains[1]
        d10: sc.Domain = s1.domains[0]
        d11: sc.Domain = s1.domains[1]
        d00.set_name("x")
        d01.set_name("y")
        d10.set_name("x")
        d11.set_name("y*")

        dsd_design = nc.Design.from_scadnano_design(sc_design)
        dsd_d00 = dsd_design.strands[0].domains[0]
        dsd_d01 = dsd_design.strands[0].domains[1]
        dsd_d10 = dsd_design.strands[1].domains[0]
        dsd_d11 = dsd_design.strands[1].domains[1]

        assert dsd_d00.name == "x"
        assert dsd_d01.name == "y"
        assert dsd_d10.name == "x"
        assert dsd_d11.name == "y"

        assert dsd_d00 is dsd_d10
        assert dsd_d01 is dsd_d11


class TestExportDNASequences:
    def test_idt_bulk_export(self) -> None:
        custom_idt = nc.VendorFields(scale="100nm", purification="PAGE")
        design = nc.Design()
        design.add_strand(domain_names=["a", "b*", "c", "d*"], name="s0", vendor_fields=custom_idt)
        design.add_strand(domain_names=["d", "c*", "e", "f"], name="s1")

        #        a       b       c       d       e           f
        seqs = ["AACG", "CCGT", "GGTA", "TTAC", "AAAACCCC", "AAAAGGGG"]
        # s0: AACG-ACGG-GGTA-GTAA
        # s1: TTAC-TACC-AAAACCCC-AAAAGGGG
        for domain, seq in zip(design.domains, seqs):
            domain.set_fixed_sequence(seq)

        idt_bulk_input = design.to_idt_bulk_input_format()
        for i, line in enumerate(idt_bulk_input.splitlines()):
            name, seq, scale, pur = line.split(",")
            if i == 0:
                assert name == "s0"
                assert seq == "AACGACGGGGTAGTAA"
                assert scale == "100nm"
                assert pur == "PAGE"
            elif i == 1:
                assert name == "s1"
                assert seq == "TTACTACCAAAACCCCAAAAGGGG"
                assert scale == "25nm"
                assert pur == "STD"

    def test_write_idt_plate_excel_file(self) -> None:
        strand_len = 10

        # add 10 strands in excess of 3 plates
        for plate_type in [sc.PlateType.wells96, sc.PlateType.wells384]:
            filename = f"test_excel_export_{plate_type.num_wells_per_plate()}.xlsx"

            design = nc.Design()
            for strand_idx in range(3 * plate_type.num_wells_per_plate() + 10):
                idt = nc.VendorFields()
                strand = design.add_strand(name=f"s{strand_idx}", domain_names=[f"d{strand_idx}"], vendor_fields=idt)
                strand.domains[0].set_fixed_sequence("T" * strand_len)

            design.write_idt_plate_excel_file(filename=filename, plate_type=plate_type)

            book = openpyxl.load_workbook(filename=filename)
            assert len(book.worksheets) == 4
            for plate in range(4):
                sheet = book.worksheets[plate]
                assert sheet.max_column == 3

                if plate == 2:  # penultimate plate
                    expected_wells = plate_type.num_wells_per_plate() - plate_type.min_wells_per_plate() + 10
                elif plate == 3:  # last plate
                    expected_wells = plate_type.min_wells_per_plate()
                else:
                    expected_wells = plate_type.num_wells_per_plate()

                assert sheet.max_row == expected_wells + 1

            os.remove(filename)


class TestNumpyFilters:
    def test_NearestNeighborEnergyFilter_raises_exception_if_energies_in_wrong_order(self) -> None:
        with pytest.raises(ValueError):
            nc.NearestNeighborEnergyFilter(-10, -15)


class TestInsertDomains:
    def setup_method(self) -> None:
        self.design = Design()
        self.design.add_strand(domain_names=["a", "b*", "c", "d*"])
        self.strand = self.design.strands[0]

    def test_no_insertion(self) -> None:
        # 0 1  2 3
        # a-b*-c-d*
        assert self.strand.starred_domain_indices == {1, 3}

    def test_append_domain_unstarred(self) -> None:
        # 0 1  2 3  4
        # a-b*-c-d*-e
        self.strand.append_domain(Domain("e"))
        assert self.strand.starred_domain_indices == {1, 3}

    def test_append_domain_starred(self) -> None:
        # 0 1  2 3  4
        # a-b*-c-d*-e*
        self.strand.append_domain(Domain("e"), starred=True)
        assert self.strand.starred_domain_indices == {1, 3, 4}

    def test_prepend_domain_unstarred(self) -> None:
        # 0 1 2  3  4
        # e-a-b*-c-d*
        self.strand.prepend_domain(Domain("e"))
        assert self.strand.starred_domain_indices == {2, 4}

    def test_prepend_domain_starred(self) -> None:
        # 0  1 2  3  4
        # e*-a-b*-c-d*
        self.strand.prepend_domain(Domain("e"), starred=True)
        assert self.strand.starred_domain_indices == {0, 2, 4}

    def test_insert_idx_2_domain_unstarred(self) -> None:
        # 0 1  2 3 4
        # a-b*-e-c-d*
        self.strand.insert_domain(2, Domain("e"))
        assert self.strand.starred_domain_indices == {1, 4}

    def test_insert_idx_2_domain_starred(self) -> None:
        # 0 1  2  3 4
        # a-b*-e*-c-d*
        self.strand.insert_domain(2, Domain("e"), starred=True)
        assert self.strand.starred_domain_indices == {1, 2, 4}


class TestExteriorBaseTypeOfDomain3PEnd:
    def test_adjacent_to_exterior_base_pair_on_length_2_domain(self) -> None:
        """Test that base pair on domain of length two is properly classified as
        ADJACENT_TO_EXTERIOR_BASE_PAIR

        .. code-block:: none

                    this base pair type should be ADJACENT_TO_EXTERIOR_BASE_PAIR
                          |
                          V
                 b        a
          <=============--==]
           |||||||||||||  ||
          [=============--==>
                 b*       a*
        """
        design = Design()
        top_strand = construct_strand(design, ["a", "b"], [2, 13])
        bot_strand = construct_strand(design, ["b*", "a*"], [13, 2])

        top_a = top_strand.address_of_domain(0)

        all_bound_domain_addresses = _get_implicitly_bound_domain_addresses([top_strand, bot_strand])

        assert (
            _exterior_base_type_of_domain_3p_end(top_a, all_bound_domain_addresses)
            == BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR
        )

    @pytest.mark.skip("MISMATCH detection has not been implemented")
    def test_mismatch(self):
        """Test MISMATCH is properly classified

        .. code-block:: none

               c    b    a
            <=====--=--=====]
             |||||     |||||
            [=====--=--=====>
               c*   d    a*
        """
        top_strand = construct_strand(["a", "b", "c"], [5, 1, 5])
        bot_strand = construct_strand(["c*", "d", "a*"], [5, 1, 5])

        top_a = top_strand.address_of_domain(0)

        all_bound_domain_addresses = _get_implicitly_bound_domain_addresses([top_strand, bot_strand])

        assert _exterior_base_type_of_domain_3p_end(top_a, all_bound_domain_addresses) == BasePairType.MISMATCH

    @pytest.mark.skip("BULGE_LOOP_3P detection has not been implemented")
    def test_bulge_loop_3p(self):
        """Test BULGE_LOOP_3P is properly classified

        .. code-block:: none

               c    b    a
            <=====--=--=====]
             |||||     |||||
            [=====-----=====>
               c*        a*
        """
        top_strand = construct_strand(["a", "b", "c"], [5, 1, 5])
        bot_strand = construct_strand(["c*", "a*"], [5, 5])

        all_bound_domain_addresses = _get_implicitly_bound_domain_addresses([top_strand, bot_strand])

        top_a = top_strand.address_of_domain(0)

        assert _exterior_base_type_of_domain_3p_end(top_a, all_bound_domain_addresses) == BasePairType.BULGE_LOOP_3P

    @pytest.mark.skip("BULGE_LOOP_5P detection has not been implemented")
    def test_bulge_loop_5p(self):
        """Test BULGE_LOOP_5P is properly classified

        .. code-block:: none

               c    b    a
            <=====--=--=====]
             |||||     |||||
            [=====-----=====>
               c*        a*
        """
        top_strand = construct_strand(["a", "c"], [5, 5])
        bot_strand = construct_strand(["c*", "d", "a*"], [5, 1, 5])

        all_bound_domain_addresses = _get_implicitly_bound_domain_addresses([top_strand, bot_strand])

        top_a = top_strand.address_of_domain(0)

        assert _exterior_base_type_of_domain_3p_end(top_a, all_bound_domain_addresses) == BasePairType.BULGE_LOOP_5P


class TestGetBasePairDomainEndpointsToCheck:
    def test_seesaw_input_gate_complex(self):
        """Test endpoints for seesaw gate input:gate complex

        .. code-block:: none

          S{input}  s{input}  T       S{gate}    s{gate}
            |           |     |          |         |
        <=============--==--=====--=============--==]
                            |||||  |||||||||||||  ||
                           [=====--=============--==--=====>
                              |          |        |     |
                              T*      S{gate}* s{gate}* T*

                       21
         34          22|20 19  15 14          2  10
         |           | ||  |   |  |           |  ||
        <=============-==--=====--=============--==]
                           |||||  |||||||||||||  ||
                          [=====--=============--==--=====>
                           |   |  |           |  ||  |   |
                           35  39 40          52 |54 55  59
                                                 53

                    DANGLE_3P INTERIOR_TO_STRAND ADJACENT_TO_EXTERIOR_BASE_PAIR
                           |      |              |
        <=============-==--=====--=============--==]
                           |||||  |||||||||||||  ||
                          [=====--=============--==--=====>
                               |              |   |
                              INTERIOR_TO_STRAND  DANGLE_3P
        """
        ssg = Domain("ssg", assign_domain_pool_of_length(13), dependent=True)
        sg = Domain("sg", assign_domain_pool_of_length(2), dependent=True)
        Sg = Domain("Sg", assign_domain_pool_of_length(15), subdomains=[sg, ssg])
        T = Domain("T", assign_domain_pool_of_length(5))
        ssi = Domain("ssi", assign_domain_pool_of_length(13), dependent=True)
        si = Domain("si", assign_domain_pool_of_length(2), dependent=True)
        Si = Domain("Si", assign_domain_pool_of_length(15), subdomains=[si, ssi])
        input_strand = Strand(domains=[Sg, T, Si], starred_domain_indices=[])
        gate_base_strand = Strand(domains=[T, Sg, T], starred_domain_indices=[0, 1, 2])
        input_gate_complex = [input_strand, gate_base_strand]

        input_t = input_strand.address_of_domain(1)
        gate_base_t = gate_base_strand.address_of_domain(0)
        nonimplicit_base_pairs = [(input_t, gate_base_t)]

        expected = set(
            [
                _BasePairDomainEndpoint(
                    domain1_5p_index=15,
                    domain2_3p_index=39,
                    domain_base_length=5,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.DANGLE_3P,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=2,
                    domain2_3p_index=52,
                    domain_base_length=13,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=0,
                    domain2_3p_index=54,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.DANGLE_3P,
                    domain1_3p_domain1_base_pair_type=BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR,
                ),
            ]
        )

        actual = _get_base_pair_domain_endpoints_to_check(input_gate_complex, nonimplicit_base_pairs)
        assert actual == expected

    def test_seesaw_gate_output_complex(self):
        """Test endpoints for seesaw gate gate:output complex

        .. code-block:: none

                   S{gate}  s{gate}  T      S{output}    s{output}
                      |        |     |          |         |
               <=============--==--=====--=============--==]
                |||||||||||||  ||  |||||
        [=====--=============--==--=====>
           |          |        |     |
           T*      S{gate}* s{gate}* T*

                               21
                34          22 |20 19  15 14          2  10
                |           |  ||  |   |  |           |  ||
               <=============--==--=====--=============--==]
                |||||||||||||  ||  |||||
        [=====--=============--==--=====>
         |   |  |           |  ||  |   |
         35  39 40          52 |54 55  59
                               53

                DANGLE_5P      INTERIOR_TO_STRAND
                |              |   |
               <=============--==--=====--=============--==]
                |||||||||||||  ||  |||||
        [=====--=============--==--=====>
                            |   |      |
               INTERIOR_TO_STRAND      DANGLE_5P
        """
        design = Design()
        output_strand = construct_strand(design, ["so", "So", "T", "sg", "Sg"], [2, 13, 5, 2, 13])
        gate_base_strand = construct_strand(design, ["T*", "Sg*", "sg*", "T*"], [5, 13, 2, 5])
        gate_output_complex = [output_strand, gate_base_strand]

        output_t = output_strand.address_of_domain(2)
        gate_base_t = gate_base_strand.address_of_domain(3)
        nonimplicit_base_pairs = [(output_t, gate_base_t)]

        expected = set(
            [
                _BasePairDomainEndpoint(
                    domain1_5p_index=22,
                    domain2_3p_index=52,
                    domain_base_length=13,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.DANGLE_5P,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=20,
                    domain2_3p_index=54,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=15,
                    domain2_3p_index=59,
                    domain_base_length=5,
                    domain1_5p_domain2_base_pair_type=BasePairType.DANGLE_5P,
                    domain1_3p_domain1_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                ),
            ]
        )

        actual = _get_base_pair_domain_endpoints_to_check(gate_output_complex, nonimplicit_base_pairs)
        assert actual == expected

    def test_seesaw_threshold_complex(self):
        """Test endpoints for seesaw threshold complex

        .. code-block:: none

                                S{gate}  s{gate}
                            14          2  10
                            |           |  ||
                           <=============--==]
                            |||||||||||||  ||
                [==--=====--=============--==>
                 ||  |   |  |           |  ||
                 15|  17  21 22          34 |36
                 16                      35
            s{input}*  T*      S{gate}*    s{gate}*

                            DANGLE_5P      ADJACENT_TO_EXTERIOR_BASE_PAIR
                            |              |
                           <=============--==]
                            |||||||||||||  ||
                [==--=====--=============--==>
                                        |   |
                       INTERIOR_TO_STRAND   BLUNT_END
        """
        design = Design()
        waste_strand = construct_strand(design, ["sg", "Sg"], [2, 13])
        threshold_base_strand = construct_strand(design, ["si*", "T*", "Sg*", "sg*"], [2, 5, 13, 2])
        threshold_complex = [waste_strand, threshold_base_strand]

        expected = set(
            [
                _BasePairDomainEndpoint(
                    domain1_5p_index=2,
                    domain2_3p_index=34,
                    domain_base_length=13,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.DANGLE_5P,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=0,
                    domain2_3p_index=36,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.BLUNT_END,
                    domain1_3p_domain1_base_pair_type=BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR,
                ),
            ]
        )

        actual = _get_base_pair_domain_endpoints_to_check(threshold_complex)
        assert actual == expected

    def test_seesaw_threshold_waste_complex(self):
        """Test endpoints for seesaw threshold waste complex

        .. code-block:: none

            S{input}  s{input}   T      S{gate}    s{gate}
                          21
             34          22|20 19  15 14          2  10
             |           | ||  |   |  |           |  ||
            <=============-==--=====--=============--==]
                           ||  |||||  |||||||||||||  ||
                          [==--=====--=============--==>
                           ||  |   |  |           |  ||
                          35|  37  41 42          54 |56
                            36                       55
                      s{input}*  T*      S{gate}*    s{gate}*

                   DANGLE_3P   INTERIOR_TO_STRAND    ADJACENT_TO_EXTERIOR_BASE_PAIR
                           |   |      |              |
            <=============-==--=====--=============--==]
                           ||  |||||  |||||||||||||  ||
                          [==--=====--=============--==>
                            |      |              |   |
                            |     INTERIOR_TO_STRAND  BLUNT_END
                            ADJACENT_TO_EXTERIOR_BASE_PAIR
        """
        ssg = Domain("ssg", assign_domain_pool_of_length(13), dependent=True)
        sg = Domain("sg", assign_domain_pool_of_length(2), dependent=True)
        Sg = Domain("Sg", assign_domain_pool_of_length(15), subdomains=[sg, ssg])
        T = Domain("T", assign_domain_pool_of_length(5))
        ssi = Domain("ssi", assign_domain_pool_of_length(13), dependent=True)
        si = Domain("si", assign_domain_pool_of_length(2), dependent=True)
        Si = Domain("Si", assign_domain_pool_of_length(15), subdomains=[si, ssi])

        input_strand = Strand(domains=[Sg, T, Si], starred_domain_indices=[])
        threshold_base_strand = Strand(domains=[si, T, Sg], starred_domain_indices=[0, 1, 2])
        threshold_waste_complex = [input_strand, threshold_base_strand]

        expected = set(
            [
                _BasePairDomainEndpoint(
                    domain1_5p_index=20,
                    domain2_3p_index=36,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR,
                    domain1_3p_domain1_base_pair_type=BasePairType.DANGLE_3P,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=15,
                    domain2_3p_index=41,
                    domain_base_length=5,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=2,
                    domain2_3p_index=54,
                    domain_base_length=13,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=0,
                    domain2_3p_index=56,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.BLUNT_END,
                    domain1_3p_domain1_base_pair_type=BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR,
                ),
            ]
        )

        actual = _get_base_pair_domain_endpoints_to_check(threshold_waste_complex)
        assert actual == expected

    def test_seesaw_reporter_complex(self):
        """Test endpoints for seesaw reporter complex

        .. code-block:: none

                      S{output}   s{output}
                    14          2  10
                    |           |  ||
                   <=============--==]
                    |||||||||||||  ||
            [=====--=============--==>
             |   |  |           |  ||
             15  19 20          32 |34
                                   33
               T*     S{output}*  s{output}*

                    DANGLE_5P      ADJACENT_TO_EXTERIOR_BASE_PAIR
                    |              |
                   <=============--==]
                    |||||||||||||  ||
            [=====--=============--==>
                                |   |
               INTERIOR_TO_STRAND   BLUNT_END
        """
        design = Design()
        waste_strand = construct_strand(design, ["so", "So"], [2, 13])
        reporter_base_strand = construct_strand(design, ["T*", "So*", "so*"], [5, 13, 2])
        reporter_complex = [waste_strand, reporter_base_strand]

        expected = set(
            [
                _BasePairDomainEndpoint(
                    domain1_5p_index=2,
                    domain2_3p_index=32,
                    domain_base_length=13,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.DANGLE_5P,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=0,
                    domain2_3p_index=34,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.BLUNT_END,
                    domain1_3p_domain1_base_pair_type=BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR,
                ),
            ]
        )

        actual = _get_base_pair_domain_endpoints_to_check(reporter_complex)
        assert actual == expected

    def test_seesaw_reporter_waste_complex(self):
        """Test endpoints for seesaw reporter waste complex

        .. code-block:: none

                S{gate}  s{gate}  T      S{output}    s{output}
                            21
             34          22 |20 19  15 14          2  10
             |           |  ||  |   |  |           |  ||
            <=============--==--=====--=============--==]
                                |||||  |||||||||||||  ||
                               [=====--=============--==>
                                |   |  |           |  ||
                                35  39 40          52 |54
                                                      53
                                  T*     S{output}*  s{output}*

                        DANGLE_3P INTERIOR_TO_STRAND  ADJACENT_TO_EXTERIOR_BASE_PAIR
                                |      |              |
            <=============--==--=====--=============--==]
                                |||||  |||||||||||||  ||
                               [=====--=============--==>
                                    |              |   |
                                   INTERIOR_TO_STRAND  BLUNT_END
        """
        design = Design()
        output_strand = construct_strand(design, ["so", "So", "T", "sg", "Sg"], [2, 13, 5, 2, 13])
        reporter_base_strand = construct_strand(design, ["T*", "So*", "so*"], [5, 13, 2])
        reporter_waste_complex = [output_strand, reporter_base_strand]

        expected = set(
            [
                _BasePairDomainEndpoint(
                    domain1_5p_index=15,
                    domain2_3p_index=39,
                    domain_base_length=5,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.DANGLE_3P,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=2,
                    domain2_3p_index=52,
                    domain_base_length=13,
                    domain1_5p_domain2_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                    domain1_3p_domain1_base_pair_type=BasePairType.INTERIOR_TO_STRAND,
                ),
                _BasePairDomainEndpoint(
                    domain1_5p_index=0,
                    domain2_3p_index=54,
                    domain_base_length=2,
                    domain1_5p_domain2_base_pair_type=BasePairType.BLUNT_END,
                    domain1_3p_domain1_base_pair_type=BasePairType.ADJACENT_TO_EXTERIOR_BASE_PAIR,
                ),
            ]
        )

        actual = _get_base_pair_domain_endpoints_to_check(reporter_waste_complex)
        assert actual == expected


class TestStrandDomainAddress:
    def setup_method(self):
        design = Design()
        self.strand = construct_strand(design, ["a", "b", "c"], [10, 20, 30])
        self.addr = StrandDomainAddress(self.strand, 1)

    def test_init(self):
        assert self.addr.strand == self.strand
        assert self.addr.domain_idx == 1

    def test_neighbor_5p(self):
        assert self.addr.neighbor_5p() == StrandDomainAddress(self.strand, 0)

    def test_neighbor_5p_none(self):
        addr = StrandDomainAddress(self.strand, 0)
        assert addr.neighbor_5p() is None

    def test_neighbor_3p(self):
        assert self.addr.neighbor_3p() == StrandDomainAddress(self.strand, 2)

    def test_neighbor_3p_none(self):
        addr = StrandDomainAddress(self.strand, 2)
        assert addr.neighbor_3p() is None

    def test_domain(self):
        assert self.addr.domain() == self.strand.domains[1]


class TestSubdomains:
    def test_init(self):
        """
        Test constructing a domain with subdomains

        .. code-block:: none

                       a
            <====================]

                 b      c      d      e
            <--=====--=====--=====--=====]
        """
        b = Domain("b", assign_domain_pool_of_length(5), dependent=True)
        c = Domain("c", assign_domain_pool_of_length(5), dependent=True)
        d = Domain("d", assign_domain_pool_of_length(5), dependent=True)
        e = Domain("e", assign_domain_pool_of_length(5), dependent=True)

        a = Domain("a", assign_domain_pool_of_length(20), subdomains=[b, c, d, e])
        assert a.subdomains == [b, c, d, e]
        assert b.parent == a
        assert c.parent == a
        assert d.parent == a
        assert e.parent == a

    def test_construct_fixed_domain_with_subdomains_should_raise_error(self):
        r"""
        Test constructing a fixed domain with fixed subdomains

        .. code-block:: none

               [a]
               / \
             [b] [c]
        """
        b = Domain("b", assign_domain_pool_of_length(5), fixed=True)
        c = Domain("c", assign_domain_pool_of_length(4), fixed=True)

        with pytest.raises(ValueError):
            _a = Domain("a", assign_domain_pool_of_length(9), fixed=True, subdomains=[b, c])

    def test_construct_unfixed_domain_with_unfixed_subdomain(self):
        r"""
        Test constructing an unfixed domain with a unfixed subdomain should
        set domain's fixed to False.

        .. code-block:: none

                a
               / \
              b  [c]
        """
        b = Domain("b", assign_domain_pool_of_length(5), fixed=False)
        c = Domain("c", assign_domain_pool_of_length(4), fixed=True)

        a = Domain("a", assign_domain_pool_of_length(9), subdomains=[b, c], fixed=False)
        assert not a.fixed

    def test_error_construct_fixed_domain_with_unfixed_subdomain(self):
        r"""
        Test that constructing a fixed domain with a unfixed subdomain should
        raise ValueError.

        .. code-block:: none

               [a]
               / \
              b  [c]
        """
        b = Domain("b", assign_domain_pool_of_length(5), fixed=False)
        c = Domain("c", assign_domain_pool_of_length(4), fixed=True)

        with pytest.raises(ValueError):
            Domain("a", assign_domain_pool_of_length(9), fixed=True, subdomains=[b, c])

    def test_error_constructed_unfixed_domain_with_fixed_subdomains(self):
        r"""
        Test that constructing a domain by setting fixed to False when all subdomains
        are fixed should raise ValueError

        .. code-block:: none

                a
               / \
             [b] [c]
        """
        b = Domain("b", assign_domain_pool_of_length(5), fixed=True)
        c = Domain("c", assign_domain_pool_of_length(4), fixed=True)

        with pytest.raises(ValueError):
            Domain("a", assign_domain_pool_of_length(9), fixed=False, subdomains=[b, c])

    def test_construct_strand(self):
        r"""
        Test strand construction with nested subdomains

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h
        """
        E = Domain("e", assign_domain_pool_of_length(5), dependent=False)
        F = Domain("f", assign_domain_pool_of_length(5), dependent=False)
        g = Domain("g", assign_domain_pool_of_length(5), dependent=True)
        h = Domain("h", assign_domain_pool_of_length(5), dependent=True)

        b = Domain("b", assign_domain_pool_of_length(10), dependent=True, subdomains=[E, F])
        C = Domain("C", assign_domain_pool_of_length(10), dependent=False, subdomains=[g, h])

        a = Domain("a", assign_domain_pool_of_length(20), dependent=True, subdomains=[b, C])

        # Test that constructor runs without errors
        strand = Strand(domains=[a], starred_domain_indices=[])
        assert strand.domains[0] == a

    def test_error_strand_with_unassignable_subsequence(self):
        r"""
        Test that constructing a strand with an unassignable subsequence raises
        a ValueError.

        This happens due to when no independent domain assigns a sequence for a
        portion of a strand

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             e   f g   h
        """
        e = Domain("e", assign_domain_pool_of_length(5), dependent=True)
        f = Domain("f", assign_domain_pool_of_length(5), dependent=True)
        g = Domain("g", assign_domain_pool_of_length(5), dependent=True)
        h = Domain("h", assign_domain_pool_of_length(5), dependent=True)

        b = Domain("b", assign_domain_pool_of_length(10), dependent=True, subdomains=[e, f])
        C = Domain("C", assign_domain_pool_of_length(10), dependent=False, subdomains=[g, h])

        a = Domain("a", assign_domain_pool_of_length(20), dependent=True, subdomains=[b, C])

        strand = Strand(domains=[a], starred_domain_indices=[])

        with pytest.raises(ValueError):
            Design(strands=[strand])

    def test_error_strand_with_redundant_independence(self):
        r"""
        Test that constructing a strand with an redundant indepndence in subdomain
        graph raises a ValueError.

        Below, in the path from F to a, two independent subdomains are found: F and B

        .. code-block:: none

                  a
                /   \
               B     C
              / \   / \
             e   F g   h
        """
        e = Domain("e", assign_domain_pool_of_length(5), dependent=True)
        F = Domain("F", assign_domain_pool_of_length(5), dependent=False)
        g = Domain("g", assign_domain_pool_of_length(5), dependent=True)
        h = Domain("h", assign_domain_pool_of_length(5), dependent=True)

        B = Domain("B", assign_domain_pool_of_length(10), dependent=False, subdomains=[e, F])
        C = Domain("C", assign_domain_pool_of_length(10), dependent=False, subdomains=[g, h])

        a = Domain("a", assign_domain_pool_of_length(20), dependent=True, subdomains=[B, C])

        strand = Strand(domains=[a], starred_domain_indices=[])

        with pytest.raises(ValueError):
            Design(strands=[strand])

    def test_error_cycle(self):
        """
        Test that constructing a domain with a cycle in its subdomain graph
        raises a ValueError.

        This isn't checked when instantiating objects, but when first calling search_for_dna_sequences,
        which calls Design.check_subdomain_graphs().

        .. code-block:: none

            a
            |
            b
            |
            a
        """
        a = Domain("a", assign_domain_pool_of_length(5), dependent=True)
        b = Domain("b", assign_domain_pool_of_length(5), subdomains=[a], dependent=True)
        a.subdomains = [b]
        strand = Strand(domains=[a], starred_domain_indices=[])

        with pytest.raises(ValueError):
            _design = Design(strands=[strand])

    def sample_nested_domains(self) -> Dict[str, Domain]:
        r"""Returns domains with the following subdomain hierarchy:

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h

        :return: Map of domain name to domain object.
        :rtype: Dict[str, Domain]
        """
        E: Domain = Domain("E", assign_domain_pool_of_length(5), dependent=False)
        F: Domain = Domain("F", assign_domain_pool_of_length(6), dependent=False)
        g: Domain = Domain("g", assign_domain_pool_of_length(7), dependent=True)
        h: Domain = Domain("h", assign_domain_pool_of_length(8), dependent=True)

        b: Domain = Domain("b", assign_domain_pool_of_length(11), dependent=True, subdomains=[E, F])
        C: Domain = Domain("C", assign_domain_pool_of_length(15), dependent=False, subdomains=[g, h])

        a: Domain = Domain("a", assign_domain_pool_of_length(26), dependent=True, subdomains=[b, C])
        return {domain.name: domain for domain in [a, b, C, E, F, g, h]}

    def test_assign_dna_sequence_to_parent(self):
        r"""
        Test assigning dna sequence to parent (a) and propagating it downwards

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h
        """
        domains = self.sample_nested_domains()
        sequence = "CATAGCTTTCTTGTTCTGATCGGAAC"
        a = domains["a"]
        a.set_sequence(sequence)
        assert a.sequence() == sequence
        assert domains["b"].sequence() == sequence[0:11]
        assert domains["C"].sequence() == sequence[11:]
        assert domains["E"].sequence() == sequence[0:5]
        assert domains["F"].sequence() == sequence[5:11]
        g_domain = domains["g"]
        assert g_domain.sequence() == sequence[11:18]
        assert domains["g"].sequence() == sequence[11:18]
        assert domains["h"].sequence() == sequence[18:]

    def test_assign_dna_sequence_to_leaf(self):
        """
        Test assigning dna sequence to E, F and propgate upward to b

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h
        """
        domains = self.sample_nested_domains()
        E = domains["E"]
        F = domains["F"]
        E.set_sequence("CATAG")
        F.set_sequence("CTTTCC")
        assert E.sequence() == "CATAG"
        assert F.sequence() == "CTTTCC"
        assert domains["b"].sequence() == "CATAGCTTTCC"

    def test_assign_dna_sequence_mixed(self):
        """
        Test assigning dna sequence to E, F, and C and propgate to entire tree.

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h
        """
        domains = self.sample_nested_domains()
        E = domains["E"]
        F = domains["F"]
        C = domains["C"]
        E.set_sequence("CATAG")
        F.set_sequence("CTTTCT")
        C.set_sequence("TGTTCTGATCGGAAC")

        # Assert initial assignment is correct
        assert domains["a"].sequence() == "CATAGCTTTCTTGTTCTGATCGGAAC"
        assert domains["b"].sequence() == "CATAGCTTTCT"
        assert domains["C"].sequence() == "TGTTCTGATCGGAAC"
        assert domains["E"].sequence() == "CATAG"
        assert domains["F"].sequence() == "CTTTCT"
        assert domains["g"].sequence() == "TGTTCTG"
        assert domains["h"].sequence() == "ATCGGAAC"

        # Assert subsequent reassignment to leaf is correct
        F.set_sequence("ATGTTT")
        assert domains["a"].sequence() == "CATAGATGTTTTGTTCTGATCGGAAC"
        assert domains["b"].sequence() == "CATAGATGTTT"
        assert domains["C"].sequence() == "TGTTCTGATCGGAAC"
        assert domains["E"].sequence() == "CATAG"
        assert domains["F"].sequence() == "ATGTTT"
        assert domains["g"].sequence() == "TGTTCTG"
        assert domains["h"].sequence() == "ATCGGAAC"

        # Assert subsequent reassignment to internal node is correct
        C.set_sequence("GGGGGGGGGGGGGGG")
        assert domains["a"].sequence() == "CATAGATGTTTGGGGGGGGGGGGGGG"
        assert domains["b"].sequence() == "CATAGATGTTT"
        assert domains["C"].sequence() == "GGGGGGGGGGGGGGG"
        assert domains["E"].sequence() == "CATAG"
        assert domains["F"].sequence() == "ATGTTT"
        assert domains["g"].sequence() == "GGGGGGG"
        assert domains["h"].sequence() == "GGGGGGGG"

    def test_error_assign_dna_sequence_to_parent_with_incorrect_size_subdomain(self):
        """
        Test error is raised if assigning dna sequence to domain when subdomains
        length do not add up to domain length.

        .. code-block:: none

                  a
                /   \
               B     C
        """
        B: Domain = Domain("B", assign_domain_pool_of_length(10), dependent=False)
        C: Domain = Domain("C", assign_domain_pool_of_length(20), dependent=False)

        a: Domain = Domain("a", assign_domain_pool_of_length(15), dependent=True, subdomains=[B, C])
        with pytest.raises(ValueError):
            a.set_sequence("A" * 15)

    def test_construct_strand_using_dependent_subdomain(self) -> None:
        """Test constructing a strand using a dependent subdomain (not parent)

        .. code-block:: none

                  a
                /   \
               b     C
              / \   / \
             E   F g   h

        Test constructing a strand using g.
        """
        g = self.sample_nested_domains()["g"]
        Strand(domains=[g], starred_domain_indices=[])

    def test_design_finds_independent_subdomains(self) -> None:
        B: Domain = Domain("B", assign_domain_pool_of_length(10), dependent=False)
        C: Domain = Domain("C", assign_domain_pool_of_length(20), dependent=False)
        a: Domain = Domain("a", assign_domain_pool_of_length(30), dependent=True, subdomains=[B, C])

        strand_a: Strand = Strand(domains=[a], starred_domain_indices=[])
        strand_b: Strand = Strand(domains=[B], starred_domain_indices=[])
        design = Design(strands=[strand_a, strand_b])
        domains = design.domains
        assert len(domains) == 3
        assert a in domains
        assert B in domains
        assert C in domains


class TestNUPACK:
    def test_pfunc(self) -> None:
        seq = "ACGTACGTAGCTGATCCAGCTGATCG"
        energy = nv.pfunc(seq)
        assert energy < 0


def _make_domain(name: str, length: int = 8, fixed: bool = False) -> Domain:
    """Helper to create a Domain with a DomainPool of a given length."""
    pool = DomainPool(f"pool_{name}", length)
    return Domain(name, pool, fixed=fixed)


def _make_result(excess: float, score: float, part: nc.Part) -> nc.Result:
    """Helper to create a Result with a specific score (bypassing score_transfer_function)."""
    result = nc.Result(excess=excess, summary=f"excess={excess}")
    result.score = score
    result.part = part
    return result


def _make_evaluation(constraint: nc.Constraint, domains: tuple[Domain, ...], result: nc.Result) -> Evaluation:
    """Helper to create an Evaluation."""
    return Evaluation(constraint=constraint, domains=domains, result=result)


def _make_domain_constraint(description: str, weight: float = 1.0) -> nc.DomainConstraint:
    """Helper to create a DomainConstraint with a dummy evaluate function."""
    return nc.DomainConstraint(
        description=description,
        short_description=description,
        weight=weight,
        evaluate=lambda seqs, part: nc.Result(excess=0.0, summary="dummy"),
    )
